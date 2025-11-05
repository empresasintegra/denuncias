# service_datatable.py - VERSIÓN FINAL CORREGIDA
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.db.models import Q, Count
from .models import Denuncia, Usuario, AdminDenuncias, Foro, Empresa
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import datetime
import os


@method_decorator(csrf_exempt, name='dispatch')
class SimpleDenunciaDataTableAPIView(APIView):
    """
    API simplificada para DataTable de denuncias - VERSIÓN CORREGIDA
    """
    
    def post(self, request, *args, **kwargs):
        try:
            print("="*50)
            print("DATATABLE API - REQUEST RECIBIDO")
            print("="*50)
            
            # ✅ FIX PRINCIPAL: Usar request.data de DRF directamente
            # DRF ya parseó el body, no intentar leerlo de nuevo
            dt_data = {}
            
            # Obtener datos de forma segura
            if hasattr(request, 'data') and request.data:
                # Si es un dict normal, usarlo
                if isinstance(request.data, dict):
                    dt_data = request.data
                # Si es QueryDict, convertir a dict normal
                else:
                    dt_data = dict(request.data)
                    # Limpiar listas de QueryDict
                    for key, value in dt_data.items():
                        if isinstance(value, list) and len(value) == 1:
                            dt_data[key] = value[0]
            
            # Asegurar valores por defecto si faltan
            dt_data.setdefault('draw', 1)
            dt_data.setdefault('start', 0)
            dt_data.setdefault('length', 10)
            dt_data.setdefault('search', {'value': '', 'regex': False})
            dt_data.setdefault('order', [{'column': 1, 'dir': 'desc'}])
            dt_data.setdefault('columns', [])
            dt_data.setdefault('user_info', {})
            
            # Convertir a enteros los valores que deben serlo
            dt_data['draw'] = int(dt_data.get('draw', 1))
            dt_data['start'] = int(dt_data.get('start', 0))
            dt_data['length'] = int(dt_data.get('length', 10))
            
            print(f"Params procesados: draw={dt_data['draw']}, start={dt_data['start']}, length={dt_data['length']}")
            print(f"User info: {dt_data.get('user_info', {})}")
            
            # ✅ FIX: Verificar autenticación antes de acceder a user
            admin = None
            if request.user and request.user.is_authenticated:
                try:
                    admin = AdminDenuncias.objects.filter(id=request.user.id).first()
                    print(f"Admin autenticado: {request.user.username} (ID: {request.user.id})")
                except Exception as e:
                    print(f"Error obteniendo admin: {e}")
                    admin = None
            else:
                print("Usuario no autenticado")

            # Construir queryset base
            denuncia = Denuncia.objects.select_related(
                'usuario',
                'item', 
                'item__categoria',
                'relacion_empresa',
                'tiempo',
                'tipo_empresa'  # Agregar esto si existe
            )
            
            # Agregar anotaciones según si hay admin o no
            if admin:
                denuncia = denuncia.annotate(
                    num_archivos=Count('archivo'),
                    num_mensajes_no_leidos=Count(
                        'foro',
                        filter=Q(foro__leido=False, foro__admin=admin)
                    ),
                    num_mensajes_leidos=Count(
                        'foro',
                        filter=Q(foro__leido=True, foro__admin=admin)
                    ),
                    num_mensajes_total=Count(
                        'foro',
                        filter=Q(foro__admin=admin)
                    )
                )
            else:
                denuncia = denuncia.annotate(
                    num_archivos=Count('archivo')
                )
            
            # FILTRADO según tipo de usuario
            if request.user and request.user.is_authenticated:
                print("Usuario autenticado - aplicando filtros de admin")
                
                # ✅ FIX: Verificar que rol_categoria existe
                if hasattr(request.user, 'rol_categoria') and request.user.rol_categoria:
                    print(f"Filtrando por categoría: {request.user.rol_categoria.nombre} (ID: {request.user.rol_categoria.id})")
                    denuncia = denuncia.filter(item__categoria_id=request.user.rol_categoria.id)
                elif request.user.is_superuser:
                    print("Superusuario - mostrando todas las denuncias")
                    # No aplicar filtros
                else:
                    print("Admin sin categoría - sin permisos")
                    denuncia = denuncia.none()
            else:
                # Usuarios no autenticados
                user_info = dt_data.get('user_info', {})
                user_type = user_info.get('tipo', 'guest')
                codigo = user_info.get('codigo', '')
                
                print(f"Usuario no autenticado - tipo: {user_type}, código: {codigo}")
                
                if user_type == 'anonimo' and codigo:
                    print(f"Filtrando denuncias anónimas: {codigo}")
                    denuncia = denuncia.filter(codigo=codigo)
                elif user_type == 'identificado' and codigo:
                    print(f"Filtrando denuncias identificadas: {codigo}")
                    denuncia = denuncia.filter(usuario__id=codigo)
                else:
                    print("Sin código válido - sin denuncias")
                    denuncia = denuncia.none()
            
            # Contar total antes de búsqueda
            records_total = denuncia.count()
            print(f"Total antes de búsqueda: {records_total}")
            
            # Aplicar búsqueda si existe
            search_value = dt_data.get('search', {}).get('value', '')
            if search_value:
                print(f"Aplicando búsqueda: '{search_value}'")
                denuncia = self._apply_search(denuncia, search_value)
            
            # Contar registros filtrados
            records_filtered = denuncia.count()
            print(f"Total después de búsqueda: {records_filtered}")
            
            # Aplicar ordenamiento
            denuncia = self._apply_ordering(denuncia, dt_data)
            
            # Paginar
            start = dt_data['start']
            length = dt_data['length']
            
            if length > 0:
                denuncias = denuncia[start:start + length]
            else:
                denuncias = denuncia  # Todas las filas si length = -1
            
            print(f"Paginación: desde {start}, cantidad {length if length > 0 else 'todas'}")
            
            # Serializar datos
            data = []
            for denuncia in denuncias:
                try:
                    row = {
                        'codigo': denuncia.codigo,
                        'fecha': denuncia.fecha.isoformat() if denuncia.fecha else '',
                        'categoria_id': denuncia.item.categoria.id if denuncia.item and denuncia.item.categoria else 0,
                        'categoria_nombre': denuncia.item.categoria.nombre if denuncia.item and denuncia.item.categoria else 'N/A',
                        'item_enunciado': denuncia.item.enunciado if denuncia.item else 'N/A',
                        'usuario_nombre': denuncia.usuario.nombre_completo if denuncia.usuario else 'N/A',
                        'usuario_anonimo': denuncia.usuario.anonimo if denuncia.usuario else False,
                        'estado_actual': denuncia.estado_actual or 'PENDIENTE',
                        'relacion_empresa': denuncia.relacion_empresa.rol if denuncia.relacion_empresa else 'N/A',
                        'tiempo': denuncia.tiempo.intervalo if denuncia.tiempo else 'N/A'
                    }
                    
                    # Agregar contadores si existen
                    if hasattr(denuncia, 'num_archivos'):
                        row['num_archivos'] = denuncia.num_archivos
                    else:
                        row['num_archivos'] = 0
                    
                    # Mensajes solo si hay admin
                    if admin:
                        row['num_mensajes_leidos'] = getattr(denuncia, 'num_mensajes_leidos', 0)
                        row['num_mensajes_no_leidos'] = getattr(denuncia, 'num_mensajes_no_leidos', 0)
                        row['num_mensajes_total'] = getattr(denuncia, 'num_mensajes_total', 0)
                    else:
                        row['num_mensajes_leidos'] = 0
                        row['num_mensajes_no_leidos'] = 0
                        row['num_mensajes_total'] = 0
                    
                    data.append(row)
                    
                except Exception as e:
                    print(f"Error serializando denuncia {denuncia.codigo}: {e}")
                    continue
            
            print(f"Serializadas {len(data)} denuncias")
            
            # Construir respuesta
            response_data = {
                'draw': dt_data['draw'],
                'recordsTotal': records_total,
                'recordsFiltered': records_filtered,
                'data': data
            }
            
            print(f"Respuesta final: draw={response_data['draw']}, total={records_total}, filtered={records_filtered}, data_count={len(data)}")
            print("="*50)
            
            return JsonResponse(response_data, safe=False)
            
        except Exception as e:
            import traceback
            print(f"ERROR en SimpleDenunciaDataTableAPIView: {str(e)}")
            print(traceback.format_exc())
            
            # Devolver respuesta válida de DataTables incluso en error
            return JsonResponse({
                'draw': 1,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': [],
                'error': str(e)
            }, status=200)  # Usar 200 para que DataTables lo procese
    
    def _apply_search(self, queryset, search_value):
        """Aplicar búsqueda global"""
        return queryset.filter(
            Q(codigo__icontains=search_value) |
            Q(usuario__nombre__icontains=search_value) |
            Q(usuario__apellidos__icontains=search_value) |
            Q(item__enunciado__icontains=search_value) |
            Q(item__categoria__nombre__icontains=search_value) |
            Q(descripcion__icontains=search_value)
        )
    
    def _apply_ordering(self, queryset, dt_data):
        """Aplicar ordenamiento"""
        # Mapeo de columnas a campos del modelo
        column_map = {
            0: 'codigo',
            1: 'fecha',
            2: 'item__categoria__nombre',
            3: 'item__enunciado',
            4: 'usuario__nombre',
            5: 'estado_actual'
        }
        
        # Obtener información de ordenamiento
        order_data = dt_data.get('order', [])
        if order_data and isinstance(order_data, list) and len(order_data) > 0:
            try:
                order = order_data[0]
                column_idx = int(order.get('column', 0))
                direction = order.get('dir', 'asc')
                
                # Obtener campo de ordenamiento
                order_field = column_map.get(column_idx, 'fecha')
                
                # Aplicar dirección
                if direction == 'desc':
                    order_field = f'-{order_field}'
                
                print(f"Ordenando por: {order_field}")
                return queryset.order_by(order_field)
            except Exception as e:
                print(f"Error aplicando ordenamiento: {e}")
        
        # Por defecto, ordenar por fecha descendente
        return queryset.order_by('-fecha')


@method_decorator(csrf_exempt, name='dispatch')
class ExportDenunciasExcelAPIView(APIView):
    """
    API para exportar denuncias a Excel
    Usa los mismos filtros que SimpleDenunciaDataTableAPIView
    """
    
    def post(self, request, *args, **kwargs):
        try:
            # Usar la misma lógica de filtrado que SimpleDenunciaDataTableAPIView
            admin = AdminDenuncias.objects.filter(id=request.user.id).first()
            
            # Obtener denuncias con los mismos filtros
            denuncias_queryset = self._get_filtered_denuncias(request, admin)
            
            # Generar Excel
            excel_file = self._generar_excel_denuncias(denuncias_queryset, request.user)
            
            return excel_file
            
        except Exception as e:
            print(f"Error en ExportDenunciasExcelAPIView: {str(e)}")
            return JsonResponse({
                'error': f'Error al generar Excel: {str(e)}'
            }, status=500)
    
    def _get_filtered_denuncias(self, request, admin):
        """Obtener denuncias con los mismos filtros que la DataTable"""
        # Parsear datos del request
        dt_data = self._parse_request(request)
        
        # Query base con anotaciones
        denuncias = Denuncia.objects.select_related(
            'usuario',
            'item', 
            'item__categoria',
            'relacion_empresa',
            'tiempo',
            'tipo_empresa'
        ).annotate(
            num_archivos=Count('archivo'),
            num_mensajes_no_leidos=Count(
                'foro',
                filter=Q(foro__leido=False, foro__admin=admin)
            ),
            num_mensajes_leidos=Count(
                'foro',
                filter=Q(foro__leido=True, foro__admin=admin)
            ),
            num_mensajes_total=Count(
                'foro',
                filter=Q(foro__admin=admin)
            )
        )
        
        # Aplicar mismos filtros que SimpleDenunciaDataTableAPIView
        if request.user.is_authenticated:
            if request.user.rol_categoria:
                denuncias = denuncias.filter(item__categoria_id=request.user.rol_categoria.id)
        else:
            user_info = dt_data.get('user_info', {})
            user_type = user_info.get('tipo', 'guest')
            codigo = user_info.get('codigo', '')
            
            if user_type == 'anonimo' and codigo:
                denuncias = denuncias.filter(codigo=codigo)
            elif user_type == 'identificado' and codigo:
                denuncias = denuncias.filter(usuario__id=codigo)
            else:
                denuncias = denuncias.none()
        
        # Aplicar búsqueda si existe
        search_value = dt_data.get('search', {}).get('value', '')
        if search_value:
            denuncias = denuncias.filter(
                Q(codigo__icontains=search_value) |
                Q(usuario__nombre__icontains=search_value) |
                Q(usuario__apellidos__icontains=search_value) |
                Q(item__enunciado__icontains=search_value) |
                Q(item__categoria__nombre__icontains=search_value) |
                Q(descripcion__icontains=search_value) |
                Q(estado_actual__icontains=search_value)
            )
        
        # Ordenar por fecha descendente
        return denuncias.order_by('-fecha')
    
    def _parse_request(self, request):
        """Parsear request de forma segura"""
        try:
            # Si viene como JSON
            if hasattr(request, 'content_type') and request.content_type == 'application/json':
                if hasattr(request, 'body') and request.body:
                    parsed_data = json.loads(request.body)
                    # ✅ VERIFICAR QUE SEA UN DICCIONARIO
                    if isinstance(parsed_data, dict):
                        return parsed_data
                    else:
                        print(f"Warning: JSON parsed data is not a dict: {type(parsed_data)}")
                        return {}
                else:
                    return {}
            # Si viene como form data
            else:
                if hasattr(request, 'POST') and request.POST:
                    post_data = request.POST.dict()
                    # ✅ VERIFICAR QUE SEA UN DICCIONARIO
                    if isinstance(post_data, dict):
                        return post_data
                    else:
                        print(f"Warning: POST.dict() is not a dict: {type(post_data)}")
                        return {}
                else:
                    return {}
        except json.JSONDecodeError as e:
            print(f"JSON decode error: {e}")
            return {}
        except Exception as e:
            print(f"Error parsing request: {e}")
            return {}  # ✅ SIEMPRE DEVOLVER DICCIONARIO
    
    def _generar_excel_denuncias(self, denuncias_queryset, user):
        """Generar archivo Excel con las denuncias"""
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Denuncias"
        
        # === Estilos ===
        titulo_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        header_font = Font(name='Calibri', size=11, bold=True)
        normal_font = Font(name='Calibri', size=10)
        
        # Colores
        title_background = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        secondary_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        alternating_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Título Principal ===
        ws['A1'] = "REPORTE DE DENUNCIAS"
        ws['A1'].font = titulo_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = title_background
        ws.merge_cells('A1:P1')  # Amplio hasta columna P
        
        # Aplicar bordes al título
        for col in range(1, 17):  # A hasta P
            cell = ws.cell(row=1, column=col)
            cell.border = thin_border
        
        # Establecer altura de fila
        ws.row_dimensions[1].height = 30
        
        # === Información del Reporte ===
        ws['A2'] = "INFORMACIÓN DEL REPORTE"
        ws['A2'].font = header_font
        ws['A2'].fill = secondary_header_fill
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:P2')
        
        # Fecha de generación
        ws['A3'] = "Fecha de Generación:"
        ws['B3'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['A4'] = "Usuario:"
        ws['B4'] = user.username if user.is_authenticated else "Consulta Pública"
        ws['A5'] = "Total de Registros:"
        ws['B5'] = denuncias_queryset.count()
        
        # Aplicar estilos a información del reporte
        for row in range(2, 6):
            for col in range(1, 17):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col == 1:  # Columna A - etiquetas
                    cell.font = header_font
                    cell.fill = secondary_header_fill
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:  # Otras columnas
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Combinar celdas para valores
        for row in range(3, 6):
            ws.merge_cells(f'B{row}:P{row}')
        
        # === Encabezados de Columnas ===
        header_row = 7
        
        headers = [
            ('A', 'Código'),
            ('B', 'Fecha'),
            ('C', 'Categoría'),
            ('D', 'Tipo de Denuncia'),
            ('E', 'Usuario'),
            ('F', 'Anónimo'),
            ('G', 'Estado'),
            ('H', 'Relación Empresa'),
            ('I', 'Tiempo Ocurrencia'),
            ('J', 'Empresa'),
            ('K', 'Archivos'),
            ('L', 'Msg. Leídos'),
            ('M', 'Msg. No Leídos'),
            ('N', 'Total Msg.'),
            ('O', 'Descripción'),
            ('P', 'Desc. Relación')
        ]
        
        # Aplicar encabezados
        for col, text in headers:
            ws[f'{col}{header_row}'] = text
            ws[f'{col}{header_row}'].font = header_font
            ws[f'{col}{header_row}'].border = thin_border
            ws[f'{col}{header_row}'].fill = header_fill
            ws[f'{col}{header_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Establecer altura de encabezado
        ws.row_dimensions[header_row].height = 25
        
        # === Datos de las Denuncias ===
        data_start_row = header_row + 1
        
        for i, denuncia in enumerate(denuncias_queryset):
            row = data_start_row + i
            
            # Aplicar fondo alternado
            row_fill = alternating_row_fill if i % 2 == 1 else None
            
            # Formatear fecha
            try:
                fecha_formateada = denuncia.fecha.strftime("%d/%m/%Y") if denuncia.fecha else ""
            except:
                fecha_formateada = str(denuncia.fecha) if denuncia.fecha else ""
            
            # Datos de la fila
            row_data = [
                ('A', denuncia.codigo),
                ('B', fecha_formateada),
                ('C', denuncia.item.categoria.nombre),
                ('D', denuncia.item.enunciado),
                ('E', denuncia.usuario.nombre_completo if not denuncia.usuario.anonimo else 'Anónimo'),
                ('F', 'Sí' if denuncia.usuario.anonimo else 'No'),
                ('G', denuncia.estado_actual),
                ('H', denuncia.relacion_empresa.rol),
                ('I', denuncia.tiempo.intervalo),
                ('J', denuncia.tipo_empresa.nombre if denuncia.tipo_empresa else ''),
                ('K', denuncia.num_archivos),
                ('L', denuncia.num_mensajes_leidos),
                ('M', denuncia.num_mensajes_no_leidos),
                ('N', denuncia.num_mensajes_total),
                ('O', denuncia.descripcion[:100] + '...' if len(denuncia.descripcion) > 100 else denuncia.descripcion),
                ('P', (denuncia.descripcion_relacion[:100] + '...' if len(denuncia.descripcion_relacion) > 100 else denuncia.descripcion_relacion) if denuncia.descripcion_relacion else '')
            ]
            
            # Aplicar datos y estilos
            for col, value in row_data:
                cell = ws[f'{col}{row}']
                cell.value = value
                cell.font = normal_font
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                
                # Alineación específica por tipo de dato
                if col in ['K', 'L', 'M', 'N']:  # Números
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col in ['O', 'P']:  # Descripciones largas
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # === Resumen Final ===
        summary_row = data_start_row + denuncias_queryset.count() + 2
        
        ws[f'A{summary_row}'] = f"Reporte generado automáticamente el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{summary_row}'].font = Font(name='Calibri', size=9, italic=True)
        ws[f'A{summary_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{summary_row}:P{summary_row}')
        
        # === Ajustar Anchos de Columnas ===
        column_widths = {
            'A': 15,  # Código
            'B': 12,  # Fecha
            'C': 20,  # Categoría
            'D': 30,  # Tipo de Denuncia
            'E': 25,  # Usuario
            'F': 10,  # Anónimo
            'G': 15,  # Estado
            'H': 20,  # Relación Empresa
            'I': 15,  # Tiempo
            'J': 15,  # Empresa
            'K': 10,  # Archivos
            'L': 12,  # Msg. Leídos
            'M': 12,  # Msg. No Leídos
            'N': 12,  # Total Msg.
            'O': 40,  # Descripción
            'P': 40   # Desc. Relación
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # === Configuración de Página ===
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Permitir múltiples páginas en altura
        
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        # === Generar nombre de archivo y guardar ===
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Reporte_Denuncias_{timestamp}.xlsx'
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el workbook directamente en la respuesta
        wb.save(response)
        
        return response


class DataTableActions(APIView):
    pass